from __future__ import annotations

import io
import re
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path

import fitz
from fastapi import FastAPI, File, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

BASE_DIR = Path(__file__).resolve().parent
TEMPLATES_DIR = BASE_DIR / "templates"
STATIC_DIR = BASE_DIR / "static"

app = FastAPI(title="ASN TOOL GM EXPORT", version="3.0.0")
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")
templates = Jinja2Templates(directory=TEMPLATES_DIR)

ASN_RE = re.compile(r"ASN No:\s*([A-Z0-9-]+)", re.I)
ETA_RE = re.compile(r"ETA:\s*([0-9-]{4,10}\s+[0-9:]{4,8})", re.I)
PACKING_RE = re.compile(r"(\d+)\*(\d+)\+(\d+)")
ROW_ONE_LINE_RE = re.compile(
    r"^(?P<seq>\d+)\s+"
    r"(?P<po>\S+)\s+"
    r"(?P<item>\d{9})\s+"
    r"(?P<rev>0[1-9])\s+"
    r"(?P<qty>\d+)\s+"
    r"(?P<uom>[A-Z]+)\s+"
    r"(?P<net>[\d.]+)\s+"
    r"(?P<pack>\d+\*\d+\+\d+)\s+"
    r"So:\s*(?P<so>\d+)\s+"
    r"(?P<lot>XC\d+)\s+"
    r"(?P<line>[A-Z]\d-[A-Z0-9-]+)$"
)
ROW_START_RE = re.compile(
    r"^(?P<seq>\d+)\s+"
    r"(?P<po>\S+)\s+"
    r"(?P<item>\d{9})\s+"
    r"(?P<rev>0[1-9])\s+"
    r"(?P<qty>\d+)\s+"
    r"(?P<uom>[A-Z]+)\s+"
    r"(?P<net>[\d.]+)\s+"
    r"(?P<pack>\d+\*\d+\+\d+)\s+"
    r"So:\s*(?P<so>\d+)$"
)
LINE_RE = re.compile(r"^[A-Z]\d-[A-Z0-9-]+$")
LOT_RE = re.compile(r"^XC\d+$")

THIN = Side(style="thin", color="C7CEDB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="1F3B6D")
SUB_FILL = PatternFill("solid", fgColor="DCE6F1")
TOTAL_FILL = PatternFill("solid", fgColor="E2F0D9")
YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


@dataclass
class ItemRow:
    seq: int
    po_no: str
    item: str
    rev: str
    quantity: int
    packing: int
    thung_chan: int
    pcs_le: int
    line_no: str
    lot_no: str = ""
    so_no: str = ""


@dataclass
class ASNData:
    asn_no: str
    eta: str
    group_name: str = "CPT"
    items: list[ItemRow] = field(default_factory=list)

    @property
    def total_quantity(self) -> int:
        return sum(i.quantity for i in self.items)

    @property
    def total_thung_chan(self) -> int:
        return sum(i.thung_chan for i in self.items)

    @property
    def total_pcs_le(self) -> int:
        return sum(i.pcs_le for i in self.items)

    @property
    def total_line_no(self) -> str:
        seen = []
        for item in self.items:
            if item.line_no not in seen:
                seen.append(item.line_no)
        return ", ".join(seen)


def clean_pdf_text(raw: str) -> list[str]:
    text = raw.replace("\u3000", " ").replace("：", ":").replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n+", "\n", text)
    return [line.strip() for line in text.split("\n") if line.strip()]


def parse_packing_spec(spec: str) -> tuple[int, int, int]:
    m = PACKING_RE.search(spec or "")
    if not m:
        return 0, 0, 0
    return int(m.group(1)), int(m.group(2)), int(m.group(3))


def infer_group(line_no: str) -> str:
    value = (line_no or "").upper()
    if value.startswith("C"):
        return "CPT"
    if value.startswith("O"):
        return "OP"
    if value.startswith("G"):
        return "GP"
    return "CPT"


def parse_delivery_note(pdf_bytes: bytes, source_name: str) -> ASNData:
    try:
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Không đọc được PDF {source_name}: {exc}") from exc

    page_text = "\n".join(page.get_text("text") for page in doc)
    doc.close()
    lines = clean_pdf_text(page_text)
    full_text = "\n".join(lines)

    asn_match = ASN_RE.search(full_text)
    eta_match = ETA_RE.search(full_text)
    if not asn_match:
        raise HTTPException(status_code=400, detail=f"Không tìm thấy ASN No trong file {source_name}")

    asn = ASNData(
        asn_no=asn_match.group(1).strip(),
        eta=eta_match.group(1).strip() if eta_match else "",
    )

    i = 0
    while i < len(lines):
        one_line = ROW_ONE_LINE_RE.match(lines[i])
        if one_line:
            packing, thung_chan, pcs_le = parse_packing_spec(one_line.group("pack"))
            asn.items.append(ItemRow(
                seq=int(one_line.group("seq")),
                po_no=one_line.group("po"),
                item=one_line.group("item"),
                rev=one_line.group("rev"),
                quantity=int(one_line.group("qty")),
                packing=packing,
                thung_chan=thung_chan,
                pcs_le=pcs_le,
                line_no=one_line.group("line"),
                lot_no=one_line.group("lot"),
                so_no=one_line.group("so"),
            ))
            i += 1
            continue

        start = ROW_START_RE.match(lines[i])
        if start and i + 2 < len(lines) and LOT_RE.match(lines[i + 1]) and LINE_RE.match(lines[i + 2]):
            packing, thung_chan, pcs_le = parse_packing_spec(start.group("pack"))
            asn.items.append(ItemRow(
                seq=int(start.group("seq")),
                po_no=start.group("po"),
                item=start.group("item"),
                rev=start.group("rev"),
                quantity=int(start.group("qty")),
                packing=packing,
                thung_chan=thung_chan,
                pcs_le=pcs_le,
                line_no=lines[i + 2],
                lot_no=lines[i + 1],
                so_no=start.group("so"),
            ))
            i += 3
            continue
        i += 1

    if not asn.items:
        raise HTTPException(status_code=400, detail=f"Không nhận dạng được dòng hàng trong file {source_name}")

    asn.group_name = infer_group(asn.items[0].line_no)
    return asn


def set_col_widths(ws) -> None:
    widths = {"A": 7, "B": 18, "C": 8, "D": 12, "E": 10, "F": 12, "G": 10, "H": 16}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def draw_title(ws, title: str) -> None:
    ws.merge_cells("A1:H1")
    cell = ws["A1"]
    cell.value = title
    cell.fill = HEADER_FILL
    cell.font = WHITE_FONT
    cell.alignment = CENTER
    cell.border = BORDER
    ws.row_dimensions[1].height = 22


def write_asn_block(ws, row: int, data: ASNData) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row, 1).value = f"ASN No: {data.asn_no}"
    for c in range(1, 9):
        ws.cell(row, c).fill = SUB_FILL
        ws.cell(row, c).border = BORDER
    ws.cell(row, 1).font = BOLD
    ws.cell(row, 1).alignment = LEFT
    row += 1

    meta = [(1, f"ETA: {data.eta}"), (3, "E ID:"), (5, "Security:"), (7, f"Nhóm: {data.group_name}")]
    for c in range(1, 9):
        ws.cell(row, c).fill = SUB_FILL
        ws.cell(row, c).border = BORDER
    for col, value in meta:
        ws.cell(row, col).value = value
        ws.cell(row, col).font = BOLD
        ws.cell(row, col).alignment = LEFT
    row += 1

    headers = ["STT", "Item", "rev", "Quantity", "Packing", "Thùng Chẵn", "PCS lẻ", "Line No."]
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row, idx)
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    row += 1

    for item in data.items:
        values = [item.seq, item.item, item.rev, item.quantity, item.packing, item.thung_chan, item.pcs_le, item.line_no]
        for idx, value in enumerate(values, start=1):
            cell = ws.cell(row, idx)
            cell.value = value
            cell.border = BORDER
            cell.alignment = CENTER if idx != 2 else LEFT
            if idx in (5, 6, 7):
                cell.fill = YELLOW_FILL
        row += 1

    totals = ["TỔNG ASN", "", "", data.total_quantity, "", data.total_thung_chan, data.total_pcs_le, data.total_line_no]
    for idx, value in enumerate(totals, start=1):
        cell = ws.cell(row, idx)
        cell.value = value
        cell.border = BORDER
        cell.fill = TOTAL_FILL
        cell.font = BOLD
        cell.alignment = CENTER
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    return row + 2


def build_workbook(asns: list[ASNData]) -> io.BytesIO:
    wb = Workbook()
    by_group: dict[str, list[ASNData]] = defaultdict(list)
    for asn in asns:
        by_group[asn.group_name].append(asn)

    for idx, group in enumerate(["CPT", "OP", "GP"]):
        ws = wb.active if idx == 0 else wb.create_sheet(group)
        ws.title = group
        set_col_widths(ws)
        draw_title(ws, f"ASN TOOL GM - {group}")
        row = 3
        if by_group[group]:
            for asn in by_group[group]:
                row = write_asn_block(ws, row, asn)
        else:
            ws.merge_cells("A3:H3")
            ws["A3"] = "Chưa có dữ liệu"
            ws["A3"].alignment = CENTER
            ws["A3"].font = BOLD

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@app.get("/", response_class=HTMLResponse)
async def home(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})


@app.get("/health")
async def health():
    return {"status": "ok"}


@app.post("/api/parse")
async def parse_only(files: list[UploadFile] = File(...)):
    results = []
    for file in files:
        data = await file.read()
        parsed = parse_delivery_note(data, file.filename or "upload.pdf")
        results.append({
            "asn_no": parsed.asn_no,
            "eta": parsed.eta,
            "group": parsed.group_name,
            "total_quantity": parsed.total_quantity,
            "total_thung_chan": parsed.total_thung_chan,
            "total_pcs_le": parsed.total_pcs_le,
            "items": [item.__dict__ for item in parsed.items],
        })
    return JSONResponse({"count": len(results), "results": results})


@app.post("/export")
async def export_excel(files: list[UploadFile] = File(...)):
    if not files:
        raise HTTPException(status_code=400, detail="Vui lòng chọn ít nhất 1 file PDF")

    parsed = []
    for file in files:
        if not (file.filename or "").lower().endswith(".pdf"):
            raise HTTPException(status_code=400, detail=f"{file.filename} không phải file PDF")
        parsed.append(parse_delivery_note(await file.read(), file.filename or "upload.pdf"))

    parsed.sort(key=lambda x: (x.group_name, x.asn_no))
    workbook = build_workbook(parsed)
    return StreamingResponse(
        workbook,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="ASN_TOOL_GM_EXPORT.xlsx"'},
    )


@app.get("/manifest.json")
async def manifest():
    return FileResponse(STATIC_DIR / "manifest.json", media_type="application/manifest+json")


@app.get("/service-worker.js")
async def service_worker():
    return FileResponse(STATIC_DIR / "service-worker.js", media_type="application/javascript")
